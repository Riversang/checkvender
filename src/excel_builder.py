"""
excel_builder.py — สร้างไฟล์ Excel ตรวจเอกสารส่วนที่ 1 และ 2

รับ list[VendorData] + project info → บันทึก .xlsx
"""
from __future__ import annotations
import os
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.page import PageMargins

from .analyzer import VendorData

# ─── สี ───────────────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="BDD7EE")
SUBHDR_FILL   = PatternFill("solid", fgColor="DDEBF7")
NEW_COL_FILL  = PatternFill("solid", fgColor="9DC3E6")  # คอลัมน์ที่แยกใหม่
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")
GRAY_FILL     = PatternFill("solid", fgColor="F2F2F2")
GREEN_FILL    = PatternFill("solid", fgColor="E2EFDA")
RED_FILL      = PatternFill("solid", fgColor="FFE0E0")
WARN_FILL     = PatternFill("solid", fgColor="FFF2CC")  # เหลือง = ขาด cert

FONT_NAME = "TH SarabunPSK"
FONT_SIZE = 14

THIN = Side(style="thin")


def _border():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _font(bold=False, size=None, color="000000"):
    return Font(name=FONT_NAME, size=size or FONT_SIZE, bold=bold, color=color)


def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _page_landscape(ws):
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(
        left=0.5, right=0.5, top=0.75, bottom=0.75, header=0.3, footer=0.3)


def _page_portrait(ws):
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(
        left=0.5, right=0.5, top=0.75, bottom=0.75, header=0.3, footer=0.3)


def _sc(ws, row, col, value, fill=None, bold=False, align=None, number_format=None):
    """Set cell shorthand"""
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(bold=bold)
    c.border = _border()
    c.alignment = align or _center()
    if fill:
        c.fill = fill
    if number_format:
        c.number_format = number_format
    return c


def _title_row(ws, text, ncols, fill=HEADER_FILL, size=16, row=1):
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws[f"A{row}"]
    c.value = text
    c.font = _font(bold=True, size=size)
    c.fill = fill
    c.alignment = _center()
    c.border = _border()


def _project_row(ws, text, ncols, row=2):
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws[f"A{row}"]
    c.value = text
    c.font = _font(size=13)
    c.fill = SUBHDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = _border()
    ws.row_dimensions[row].height = 36


def _summary_row(ws, ncols, count, row):
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws.cell(row=row, column=1, value=f"จำนวนผู้ยื่นข้อเสนอ {count} ราย")
    c.font = _font(bold=True)
    c.fill = SUBHDR_FILL
    c.alignment = _center()
    c.border = _border()


# ─── Sheet: เอกสารส่วนที่ 2 ─────────────────────────────────────────────────

_HEADERS_2 = [
    "ลำดับ", "ผู้ยื่นข้อเสนอ",
    "หนังสือ\nมอบอำนาจ", "หลักประกัน",
    "SMEs", "Made in\nThailand",
    "แคตตาล็อก/\nคุณลักษณะเฉพาะ",
    "หนังสือรับรอง\nผลงาน",                         # H
    "หนังสือรับรอง\nCertificate/\nLicense",          # I (generic — รองรับทุก project)
    "บุคลากร",                                       # J (MA / จ้าง)
    "แผน/โครงสร้าง\nการบริหาร",                     # K (MA / จ้าง)
    "หมายเหตุ",
]
_WIDTHS_2 = [6, 28, 10, 10, 8, 10, 16, 14, 18, 10, 14, 22]


def _build_sheet2(wb, vendors: list[VendorData], project_name: str):
    ws = wb.active
    ws.title = "เอกสารส่วนที่2"
    ncols = len(_HEADERS_2)

    _title_row(ws, "ตรวจเอกสารส่วนที่ 2", ncols)
    _project_row(ws, project_name, ncols)

    ws.row_dimensions[3].height = 56
    for col, h in enumerate(_HEADERS_2, 1):
        c = _sc(ws, 3, col, h, fill=HEADER_FILL, bold=True)
        if col in (8, 9):   # คอลัมน์แยกใหม่ → สีน้ำเงินเข้มขึ้น
            c.fill = NEW_COL_FILL

    for i, w in enumerate(_WIDTHS_2, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for i, v in enumerate(vendors):
        row = 4 + i
        ws.row_dimensions[row].height = 32
        fill = WHITE_FILL if i % 2 == 0 else GRAY_FILL
        wc_fill = fill if v.work_cert    == "✓" else WARN_FILL
        ll_fill = fill if v.line_license == "✓" else WARN_FILL

        _sc(ws, row, 1,  v.no,           fill=fill, bold=True)
        _sc(ws, row, 2,  v.name,         fill=fill, align=_left())
        _sc(ws, row, 3,  v.poa,          fill=fill)
        _sc(ws, row, 4,  v.guarantee,    fill=fill)
        _sc(ws, row, 5,  v.sme,          fill=fill)
        _sc(ws, row, 6,  v.mit,          fill=fill)
        _sc(ws, row, 7,  v.catalogue,    fill=fill)
        _sc(ws, row, 8,  v.work_cert,    fill=wc_fill)
        _sc(ws, row, 9,  v.line_license, fill=ll_fill)
        _sc(ws, row, 10, v.personnel,    fill=fill)
        _sc(ws, row, 11, v.project_mgmt, fill=fill)
        # หมายเหตุ: รวม other1_note + warning ไฟล์อ่านไม่ออก
        note_parts = []
        if v.other1_note:
            note_parts.append(v.other1_note)
        if v.unread_files:
            note_parts.append("⚠ อ่านไม่ออก (ต้อง OCR): "
                              + ", ".join(v.unread_files[:3])
                              + (f" +{len(v.unread_files)-3}"
                                 if len(v.unread_files) > 3 else ""))
        note_fill = WARN_FILL if v.unread_files else fill
        _sc(ws, row, 12, "\n".join(note_parts),
            fill=note_fill, align=_left())

    _summary_row(ws, ncols, len(vendors), 4 + len(vendors))
    _page_landscape(ws)


# ─── Sheet: เอกสารส่วนที่ 1 ─────────────────────────────────────────────────

_HEADERS_1 = [
    "ลำดับ", "ผู้ยื่นข้อเสนอ",
    "หนังสือ\nรับรอง", "บริคณห์\nสนธิ",
    "บัญชีกรรมการ\nผู้จัดการ", "ผู้มีอำนาจ\nควบคุม", "ผู้ถือหุ้น\nรายใหญ่",
    "ทะเบียน\nการค้า", "ภพ.20", "มูลค่าสุทธิ\n(บวก/ลบ)",
    "หนังสือรับรอง\nวงเงินสินเชื่อ",
    "สัญญา\nร่วมค้า", "ข้อตกลง\nคุณธรรม", "นโยบายป้องกัน\nทุจริต", "หมายเหตุ",
]
_WIDTHS_1 = [6, 28, 10, 10, 14, 14, 18, 10, 8, 14, 14, 10, 12, 14, 18]


def _build_sheet1(wb, vendors: list[VendorData], project_name: str):
    ws = wb.create_sheet("เอกสารส่วนที่ 1")
    ncols = len(_HEADERS_1)

    _title_row(ws, "ตรวจเอกสารส่วนที่ 1", ncols)
    _project_row(ws, project_name, ncols)

    ws.row_dimensions[3].height = 52
    for col, h in enumerate(_HEADERS_1, 1):
        _sc(ws, 3, col, h, fill=HEADER_FILL, bold=True)

    for i, w in enumerate(_WIDTHS_1, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for i, v in enumerate(vendors):
        row = 4 + i
        ws.row_dimensions[row].height = 32
        fill = WHITE_FILL if i % 2 == 0 else GRAY_FILL
        net_str = (f"{v.net_worth:,.2f}\n({v.net_worth_sign})"
                   if v.net_worth else f"({v.net_worth_sign})")

        _sc(ws, row, 1,  v.no,               fill=fill, bold=True)
        _sc(ws, row, 2,  v.name,             fill=fill, align=_left())
        _sc(ws, row, 3,  v.cert,             fill=fill)
        _sc(ws, row, 4,  v.memo,             fill=fill)
        _sc(ws, row, 5,  v.director_list,    fill=fill)
        _sc(ws, row, 6,  v.authority_doc,    fill=fill)
        _sc(ws, row, 7,  v.shareholder_doc,  fill=fill)
        _sc(ws, row, 8,  v.trade_reg,        fill=fill)
        _sc(ws, row, 9,  v.vat,              fill=fill)
        _sc(ws, row, 10, net_str,            fill=fill)
        _sc(ws, row, 11, v.credit,           fill=fill)
        _sc(ws, row, 12, v.joint,            fill=fill)
        _sc(ws, row, 13, v.integrity,        fill=fill)
        _sc(ws, row, 14, v.anti_corrupt,     fill=fill)
        # หมายเหตุ: warning ไฟล์อ่านไม่ออก
        if v.unread_files:
            note = ("⚠ อ่านไม่ออก (ต้อง OCR): "
                    + ", ".join(v.unread_files[:3])
                    + (f" +{len(v.unread_files)-3}"
                       if len(v.unread_files) > 3 else ""))
            _sc(ws, row, 15, note, fill=WARN_FILL, align=_left())
        else:
            _sc(ws, row, 15, "", fill=fill)

    _summary_row(ws, ncols, len(vendors), 4 + len(vendors))
    _page_portrait(ws)


# ─── Sheet: กรรมการ ──────────────────────────────────────────────────────────

def _build_directors(wb, vendors: list[VendorData], project_name: str):
    ws = wb.create_sheet("กรรมการ")
    ncols = 7

    _title_row(ws, "รายชื่อกรรมการ / ผู้มีอำนาจควบคุม / ผู้ถือหุ้นรายใหญ่", ncols)
    _project_row(ws, project_name, ncols)

    headers = ["ลำดับ","ชื่อบริษัท","เลขทะเบียน","จำนวน\nกรรมการ",
               "รายชื่อกรรมการ","รายชื่อผู้มีอำนาจควบคุม\n(ถ้ามี)",
               "ผู้ถือหุ้นรายใหญ่\n(>25%)"]
    ws.row_dimensions[3].height = 52
    for col, h in enumerate(headers, 1):
        _sc(ws, 3, col, h, fill=HEADER_FILL, bold=True)

    for i, w in enumerate([6, 28, 16, 10, 35, 32, 35], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for i, v in enumerate(vendors):
        row = 4 + i
        n_dir = len(v.directors)
        ws.row_dimensions[row].height = max(32, n_dir * 20)
        fill = WHITE_FILL if i % 2 == 0 else GRAY_FILL
        dirs_str = "\n".join(f"{j+1}. {d}" for j, d in enumerate(v.directors))

        _sc(ws, row, 1, v.no,       fill=fill, bold=True)
        _sc(ws, row, 2, v.name,     fill=fill, align=_left())
        _sc(ws, row, 3, v.tax_id,   fill=fill)
        _sc(ws, row, 4, n_dir,      fill=fill)
        _sc(ws, row, 5, dirs_str,   fill=fill, align=_left())
        _sc(ws, row, 6, v.authority,fill=fill, align=_left())
        _sc(ws, row, 7, v.shareholders, fill=fill, align=_left())

    _summary_row(ws, ncols, len(vendors), 4 + len(vendors))
    _page_portrait(ws)


# ─── Sheet: ราคา ─────────────────────────────────────────────────────────────

def _build_prices(wb, vendors: list[VendorData], project_name: str, budget: float):
    ws = wb.create_sheet("ราคา")

    _title_row(ws, "ตารางราคาเสนอ", 4)
    _project_row(ws, project_name, 4)
    ws.row_dimensions[2].height = 48

    # วงเงิน
    _sc(ws, 3, 1, "วงเงินงบประมาณ", fill=SUBHDR_FILL, bold=True)
    ws.merge_cells("B3:D3")
    c = ws.cell(row=3, column=2, value=budget)
    c.font = _font(bold=True); c.fill = SUBHDR_FILL
    c.border = _border(); c.alignment = _center()
    c.number_format = "#,##0.00"

    for col, h in enumerate(["ลำดับ","ผู้ยื่นข้อเสนอ","ราคาที่เสนอ (บาท)","ส่วนต่างจากวงเงิน"], 1):
        _sc(ws, 4, col, h, fill=HEADER_FILL, bold=True)
    ws.row_dimensions[4].height = 36

    for i, w in enumerate([8, 35, 20, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for i, v in enumerate(sorted(vendors, key=lambda x: x.price)):
        row = 5 + i
        ws.row_dimensions[row].height = 28
        diff = v.price - budget
        fill = GREEN_FILL if v.price <= budget else RED_FILL

        _sc(ws, row, 1, v.no,    fill=fill, bold=True)
        _sc(ws, row, 2, v.name,  fill=fill, align=_left())
        c3 = ws.cell(row=row, column=3, value=v.price)
        c3.font = _font(); c3.fill = fill; c3.border = _border()
        c3.alignment = _center(); c3.number_format = "#,##0.00"
        c4 = ws.cell(row=row, column=4, value=diff)
        c4.font = _font(); c4.fill = fill; c4.border = _border()
        c4.alignment = _center(); c4.number_format = "+#,##0.00;-#,##0.00;0.00"

    note_row = 5 + len(vendors) + 1
    ws.merge_cells(f"A{note_row}:D{note_row}")
    c = ws.cell(row=note_row, column=1,
                value="หมายเหตุ: สีเขียว = ไม่เกินวงเงิน  สีแดง = เกินวงเงิน")
    c.font = _font(size=12); c.fill = WHITE_FILL
    c.border = _border(); c.alignment = _left()

    _page_portrait(ws)


# ─── Public API ───────────────────────────────────────────────────────────────

def build_excel(
    vendors: list[VendorData],
    project_name: str,
    budget: float,
    output_path: str,
) -> str:
    """
    สร้างไฟล์ Excel และบันทึก

    Parameters
    ----------
    vendors      : ข้อมูลที่ analyze แล้ว
    project_name : ชื่อโครงการเต็ม
    budget       : วงเงินงบประมาณ (บาท)
    output_path  : path ที่บันทึก เช่น output/result.xlsx

    Returns
    -------
    str : path ที่บันทึกจริง
    """
    wb = openpyxl.Workbook()
    _build_sheet2(wb, vendors, project_name)
    _build_sheet1(wb, vendors, project_name)
    _build_directors(wb, vendors, project_name)
    _build_prices(wb, vendors, project_name, budget)

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    return output_path
